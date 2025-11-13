import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import os, sys

# ✅ Detect base directory (works after packaging with PyInstaller)
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

DB_PATH = os.path.join(BASE_DIR, "database", "farmers.db")
EXPORTS_DIR = os.path.join(BASE_DIR, "exports")

# ✅ Auto-create folders if missing
os.makedirs(os.path.join(BASE_DIR, "database"), exist_ok=True)
os.makedirs(EXPORTS_DIR, exist_ok=True)

print("📂 Using database file:", DB_PATH)
print("📁 Export directory:", EXPORTS_DIR)


def export_payments_to_excel():
    """Export summarized payments per farmer to Excel with formatting."""
    try:
        conn = sqlite3.connect(DB_PATH)

        # Summarized query per farmer
        query = """
        SELECT 
            f.name AS FarmerName,
            f.village AS Village,
            ROUND(SUM(p.total_weight), 2) AS TotalWeight_Ton,
            ROUND(AVG(p.rate_per_ton), 2) AS RatePerTon,
            ROUND(SUM(p.deductions), 2) AS TotalDeductions,
            ROUND(SUM(p.total_amount), 2) AS TotalAmount,
            MAX(p.payment_date) AS PaymentDate
        FROM payments p
        JOIN farmers f ON p.farmer_id = f.farmer_id
        GROUP BY f.farmer_id
        ORDER BY f.name
        """

        df = pd.read_sql_query(query, conn)
        conn.close()

        if df.empty:
            print("⚠️ No payment data found to export.")
            return

        # File path for Excel export
        file_path = os.path.join(EXPORTS_DIR, "payments_report.xlsx")
        df.to_excel(file_path, index=False)

        # Load and format Excel
        wb = load_workbook(file_path)
        ws = wb.active

        # Styles
        bold = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # Format header row
        for cell in ws[1]:
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                if ws.cell(row=1, column=cell.column).value == "TotalAmount":
                    cell.number_format = '"₹"#,##0.00'

        # Add total row
        total_row = ws.max_row + 2
        ws[f"E{total_row}"] = "Grand Total:"
        ws[f"F{total_row}"] = f"=SUM(F2:F{ws.max_row-1})"
        ws[f"E{total_row}"].font = Font(bold=True)
        ws[f"F{total_row}"].number_format = '"₹"#,##0.00'

        # Highlight total row
        total_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        ws[f"E{total_row}"].fill = total_fill
        ws[f"F{total_row}"].fill = total_fill

        wb.save(file_path)
        print(f"✅ Excel report exported successfully: {file_path}")

    except Exception as e:
        print("❌ Error exporting payments:", e)
